from flask import Blueprint, render_template, redirect, url_for, request, flash, session
from flask_login import login_user, login_required, logout_user, current_user
from models import User, Student, Grade
from werkzeug.utils import secure_filename
import openpyxl
from database import db_session
from functools import wraps
from random import *
import string


admin_bp = Blueprint("admin_bp", __name__, template_folder="templates", static_folder="static")

ALLOWED_EXTENSIONS = {'xlsx', 'xlsm', 'xltx', 'xltm'}


def check_admin(func):
    @wraps(func)
    def inner(*args, **kwargs):
        if "permit" in session:
            if session["permit"] == 1:
                return func(*args, **kwargs)
        return redirect(url_for("student_bp.login_page"))
    return inner


@admin_bp.route("/all/")
@check_admin
def landing_page():
    grades = Grade.query.all()
    students = Student.query.all()
    students_w_book = Student.query.filter(Student.books != None).all()
    return render_template("admin/landing_page.html", grades=grades, students=students, stac=len(students), stwbc=len(students_w_book))


@admin_bp.route("/grade/<grade>/")
@check_admin
def landing_page_grade(grade):
    cer_grade = Grade.query.filter(Grade.name == grade).first()

    if cer_grade:
        stwbc = Student.query.filter(Student.grade == cer_grade, Student.books != None).all()
        return render_template("admin/class_page.html", grade=cer_grade, students=cer_grade.students, stwbc=len(stwbc))
    flash(f"Trieda {grade} neexistuje","danger")
    return redirect(url_for("admin_bp.landing_page"))


@admin_bp.route("/list/")
@check_admin
def student_list():
    students = Student.query.all()
    return render_template("admin/student_list.html", students=students)


@admin_bp.route("/new_student/", methods=["POST"])
@check_admin
def new_student():
    name = request.form["student-name"]
    s_grade = request.form["student-grade"]
    code = create_number()
    grade = Grade.query.filter(Grade.name == s_grade).first()
    code_student = Student.query.filter(Student.code == code).first()
    if code_student:
        flash(f"Žiak s kódom {code} už existuje","danger")
        return redirect(url_for("admin_bp.landing_page"))
    if grade:
        n_student = Student(name,grade,int(code))
        db_session.add(n_student)
        db_session.commit()
        flash(f"Študent {n_student.name} bol úspešne pridaný", "success")
        return redirect(url_for("admin_bp.landing_page_grade", grade=grade.name))
    flash(f"Trieda {s_grade} neexistuje", "danger")
    return redirect(url_for("admin_bp.landing_page_grade", grade=grade.name))


@admin_bp.route("/move_all_students_up/")
@check_admin
def move_all_s_up():

    grades = ["Prima", "Sekunda", "Tercia", "Kvarta", "Kvinta", "Sexta", "Septima", "Oktava",
              "1.A","2.A","3.A","4.A",
              "1.B", "2.B", "3.B","4.B"
              ]
    final_grades = ["Oktava","4.A","4.B"]
    students = Student.query.all()

    for student in students:
        if student.grade.name in final_grades:
            db_session.delete(student)
            db_session.commit()
            continue
        index = grades.index(student.grade.name)
        new_grade = Grade.query.filter(Grade.name == grades[index+1]).first()
        student.grade = new_grade
        db_session.commit()

    flash("Všetci študenti prešli do vyšieho ročníka", "success")
    return redirect(url_for("admin_bp.student_list"))


@admin_bp.route("/move_all_students_down/")
@check_admin
def move_all_s_down():

    grades = ["Prima", "Sekunda", "Tercia", "Kvarta", "Kvinta", "Sexta", "Septima", "Oktava",
              "1.A","2.A","3.A","4.A",
              "1.B", "2.B", "3.B","4.B"
              ]
    final_grades = ["Prima","1.A","1.B"]
    students = Student.query.all()

    for student in students:
        if student.grade.name in final_grades:
            db_session.delete(student)
            db_session.commit()
            continue
        index = grades.index(student.grade.name)
        new_grade = Grade.query.filter(Grade.name == grades[index-1]).first()
        student.grade = new_grade
        db_session.commit()

    flash("Všetci študenti prešli do nižšieho ročníka", "success")
    return redirect(url_for("admin_bp.student_list"))


@admin_bp.route("/change_grade/for/<int:student_id>/",methods=["POST"])
@check_admin
def change_student(student_id):
    new_name = request.form["student_name"]
    new_code = request.form["student_code"]
    new_grade = Grade.query.filter(Grade.name == request.form["student_grade"]).first()

    student = Student.query.filter(Student.id == student_id).first()
    if not student:
        flash("Neplatný kód študenta","danger")
        return redirect(url_for("admin_bp.landing_page"))
    code_student = Student.query.filter(Student.code == new_code).first()

    if code_student and code_student.id != student.id:
        flash("Študent s takýmto kódom už existuje","danger")
        return redirect(url_for("admin_bp.view_student",student_id=student_id))

    student.name = new_name
    student.code = new_code
    student.grade = new_grade
    db_session.commit()

    flash(f"Údaje študenta {student.name} boli úspešne zmenené", "success")
    return redirect(url_for("admin_bp.view_student", student_id=student_id))


@admin_bp.route("/<int:student_id>/return_book/<int:book_id>/")
@check_admin
def return_book(student_id,book_id):

    book = Book.query.filter(Book.id == book_id).first()
    student = Student.query.filter(Student.id == student_id).first()

    if book and student:
        student.books.remove(book)
        db_session.commit()
        flash("Učebnica bola vrátená","success")

        return redirect(url_for("admin_bp.view_student",student_id=student_id))
    else:
        return redirect(url_for("admin_bp.landing_page"))


@admin_bp.route("/<int:student_id>/return_all/")
@check_admin
def return_all(student_id):
    student = Student.query.filter(Student.id == student_id).first()
    if not student:
        return redirect(url_for("admin_bp.landing_page"))

    student.books.clear()
    db_session.commit()
    flash(f"Všetky učebnice žiaka {student.name} boli úspešne vrátené","success")

    return redirect(url_for("student_bp.view_student",student_id=student_id))


@admin_bp.route("/delete/<int:student_id>/")
@check_admin
def delete_student(student_id):
    student = Student.query.filter(Student.id == student_id).first()
    grade = student.grade
    if not student:
        return redirect(url_for("admin_bp.landing_page"))
    db_session.delete(student)
    db_session.commit()
    flash(f"Študent {student.name} bol úspešne vymazaný","success")
    return redirect(url_for("admin_bp.landing_page_grade", grade=grade.name))


def create_number():
    number = int(''.join(choices(string.digits, k=8)))
    student = Student.query.filter(Student.code == number).first()
    if student:
        return create_number()
    else:
        return number


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@admin_bp.route('/upload/source/excel/', methods=["POST"])
@check_admin
def upload_file():

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active

        grade_name = request.form["grade"]
        start_point = int(request.form["start_point"])
        name_column = int(request.form["name_column"])

        try:
            for i in range(1, sheet_obj.max_row):
                s_name = sheet_obj.cell(row=i+start_point-1, column=name_column).value

                if s_name:
                    grade = Grade.query.filter(Grade.name == grade_name).first()
                    code = create_number()

                    n_student = Student(s_name, grade, code)
                    db_session.add(n_student)
            db_session.commit()
        except:
            flash("Nastala chyba pri nahrávaní. Ujistite sa, či študenti z tabuľky nie su už v systéme", "danger")
            return redirect(url_for("admin_bp.landing_page"))

        flash("Študenti z execelu boli úspešne pridaný","success")
        return redirect(url_for("admin_bp.landing_page_grade", grade=grade_name))

    allow_f_string = ' / '.join(map(str, ALLOWED_EXTENSIONS))
    flash(f"Súbor nie je podporovaný. Typ súboru musí byť: {allow_f_string}","danger")
    return redirect(url_for("admin_bp.landing_page"))
