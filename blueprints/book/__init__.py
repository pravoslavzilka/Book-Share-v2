from flask import Blueprint, render_template, redirect, url_for, request, flash, session
from database import db_session
from models import Grade, Student, Book, BookType
from werkzeug.utils import secure_filename
from flask_login import login_required
import openpyxl
from functools import wraps


book_bp = Blueprint("book_bp", __name__, template_folder="templates")

ALLOWED_EXTENSIONS = {'xlsx', 'xlsm', 'xltx', 'xltm'}


def check_admin(func):
    @wraps(func)
    def inner(*args, **kwargs):
        if "permit" in session:
            if session["permit"] == 1:
                return func(*args, **kwargs)
        return redirect(url_for("student_bp.login_page"))
    return inner


@book_bp.route("/all/")
@check_admin
def landing_page():
    books = Book.query.all()
    free_books = Book.query.filter(Book.student == None).count()
    rent_books = Book.query.filter(Book.student).count()
    total_books = Book.query.count()

    book_types = BookType.query.all()

    return render_template("book/landing_page.html", books=books, book_types=book_types,
                           ctbooks=total_books, cfbooks=free_books, crbooks=rent_books
                           )


@book_bp.route("/type/<bt_name>/")
@check_admin
def book_type_page(bt_name):
    book_type = BookType.query.filter(BookType.name == bt_name).first()

    free_books = Book.query.filter(Book.student == None, Book.book_type == book_type).count()
    rent_books = Book.query.filter(Book.student, Book.book_type == book_type).count()
    total_books = Book.query.filter(Book.book_type == book_type).count()

    return render_template("book/book_type_page.html",book_type=book_type,
                           ctbooks=total_books, cfbooks=free_books, crbooks=rent_books
                           )


@book_bp.route("/total_list/")
@check_admin
def list_of_books():
    books = Book.query.all()
    book_types = BookType.query.all()
    return render_template("book/book_list.html", books=books,book_types=book_types)


@book_bp.route("/list/<book_state>/")
@check_admin
def books_state(book_state):
    book_types = BookType.query.all()
    if book_state == "free":
        books = Book.query.filter(Book.student_id == None).all()
    else:
        books = Book.query.filter(Book.student_id != None).all()

    return render_template("book/book_list.html",books=books,book_types=book_types)


@book_bp.route("/add_new/in/<bt_name>/",methods=["POST"])
@check_admin
def add_book_with_type(bt_name):
    book_code = request.form["book_code"]
    book = Book.query.filter(Book.code == book_code).first()

    if book:
        flash("Učebnica s týmto kódom už existuje", "danger")
        return redirect(url_for("book_bp.book_type_page",bt_name=bt_name))

    book_type = BookType.query.filter(BookType.name == bt_name).first()
    new_book = Book(int(book_code), book_type)

    db_session.add(new_book)
    try:
        db_session.commit()
    except OverflowError:
        flash("Neplatný kód", "danger")
        return redirect(url_for("book_bp.book_type_page", bt_name=bt_name))

    flash(f"Učebnica s kódom {book_code} úspešne pridaná","success")
    return redirect(url_for("book_bp.book_type_page", bt_name=bt_name))


@book_bp.route("/return/",methods=["GET"])
@check_admin
def return_book_view():
    return render_template("book/return_book.html")


@book_bp.route("/return-book-s-page/<int:book_code>")
@check_admin
def return_book_s_page(book_code):
    book = Book.query.filter(Book.code == int(book_code)).first()
    if book:
        if book.student:
            student = Student.query.filter(Student.id == book.student.id).first()
            student.books.remove(book)
            db_session.commit()
            flash(f"Kniha s kódom {book.code} bola úspešne vrátená", "success")
            return redirect(url_for("student_bp.view_student", student_id=student.id))

    return redirect(url_for("student_bp.landing_page"))


@book_bp.route("/return/",methods=["POST"])
@check_admin
def return_book():
    code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == code).first()
    except OverflowError:
        flash("Neplatný kód", "danger")
        return redirect(url_for("book_bp.return_book_view"))

    if book:
        if book.student:
            student = Student.query.filter(Student.id == book.student.id).first()
            student.books.remove(book)
            db_session.commit()
        flash(f"Kniha s kódom {book.code} bola úspešne vrátená","success")
        return redirect(url_for("book_bp.return_book_view"))

    flash("Neplatný kód","danger")
    return redirect(url_for("book_bp.return_book_view"))


@book_bp.route("/return/from/list/",methods=["POST"])
@check_admin
def return_book_from_list():
    code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == code).first()
    except OverflowError:
        flash("Neplatný kód", "danger")
        return redirect(url_for("book_bp.list_of_books"))

    if book:
        if book.student:
            student = Student.query.filter(Student.id == book.student.id).first()
            student.books.remove(book)
            db_session.commit()
        flash(f"Učebnica s kódom {book.code} bola úspešne vrátená", "success")
        return redirect(url_for("book_bp.list_of_books"))

    flash("Neplatný kód", "danger")
    return redirect(url_for("book_bp.list_of_books"))


@book_bp.route("/return/with/type/",methods=["POST"])
@check_admin
def return_book_with_type():
    code = int(request.form["book_code"])
    try:
        book = Book.query.filter(Book.code == code).first()
    except OverflowError:
        flash("Neplatný kód", "danger")
        return redirect(url_for("book_bp.book_type_page"))

    if book:
        if book.student:
            student = Student.query.filter(Student.id == book.student.id).first()
            student.books.remove(book)
            db_session.commit()
        flash(f"Učebnica s kódom {book.code} bola úspešne vrátená", "success")
        return redirect(url_for("book_bp.book_type_page",bt_name=book.book_type.name))

    flash("Neplatný kód", "danger")
    return redirect(url_for("book_bp.book_type_page",bt_name=book.book_type.name))


@book_bp.route("/delete/", methods=['POST'])
@check_admin
def delete_book():
    book_id = int(request.form["book_id"])
    try:
        book = Book.query.filter(Book.id == book_id).first()
    except OverflowError:
        flash("Neplatný kód", "danger")
        return redirect(url_for("book_bp.list_of_books"))

    if book:
        book_type = BookType.query.filter(BookType.id == book.book_type_id).first()
        db_session.delete(book)
        db_session.commit()
        flash(f"Kniha s kódom {book.code} bola úspešne odstránená","success")
        return redirect(url_for("book_bp.book_type_page", bt_name=book_type.name))

    flash("Neplatný kód", "danger")
    return redirect(url_for("book_bp.list_of_books"))


@book_bp.route("/delete/from/list/", methods=['POST'])
@check_admin
def delete_book_from_list():
    print(request.form["book_id"])
    book_id = int(request.form["book_id"])
    try:
        book = Book.query.filter(Book.id == book_id).first()
    except OverflowError:
        flash("Neplatný kód","danger")
        return redirect(url_for("book_bp.list_of_books"))

    db_session.delete(book)
    db_session.commit()
    flash(f"Kniha s kódom {book.code} bola úspešne odstránená","success")
    return redirect(url_for("book_bp.list_of_books"))


@book_bp.route("/add_type/",methods=["POST"])
@check_admin
def add_book_type():
    req_bt = request.form["book_name"]
    req_ba = request.form["book_author"]

    try:
        book_type = BookType.query.filter(BookType.name == req_bt).first()
    except OverflowError:
        flash("Neplatný kód","danger")
        return redirect(url_for("book_bp.landing_page"))

    if book_type:
        flash("Typ tejto učibnice už existuje","danger")
        return redirect(url_for("book_bp.landing_page"))

    new_bt = BookType(req_bt,req_ba)
    db_session.add(new_bt)
    try:
        db_session.commit()
    except OverflowError:
        flash("Neplatné údaje", "danger")
        return redirect(url_for("book_bp.landing_page"))

    flash(f"Nový typ učebníc {req_bt} bol úspešne pridaný", "success")
    return redirect(url_for("book_bp.landing_page"))


@book_bp.route("/update/type/<int:type_id>/", methods=["POST"])
@check_admin
def update_book_type(type_id):
    book_type = BookType.query.filter(BookType.id == type_id).first()

    if book_type:
        n_name = request.form["book_type_name"]
        n_author = request.form["book_type_author"]
        book_type.name = n_name
        book_type.author = n_author
        db_session.commit()
        flash(f"Učebnica '{book_type.name}' bola aktualizovaná.","success")
        return redirect(url_for("book_bp.book_type_page", bt_name=book_type.name))
    else:
        return redirect(url_for("book_bp.landing_page"))


@book_bp.route("/delete/type/<int:type_id>/")
@check_admin
def delete_type(type_id):

    try:
        book_type = BookType.query.filter(BookType.id == type_id).first()
    except OverflowError:
        flash("Neplataný kód", "danger")
        return redirect(url_for("book_bp.landing_page"))

    count_of_books = 0
    if book_type:
        for index,book in enumerate(book_type.books):
            db_session.delete(book)
            count_of_books += 1

        db_session.delete(book_type)
        db_session.commit()
        flash(f"Typ učebnice '{book_type.name}' bol úspešne vymazaný a s ním {count_of_books} učebníc","success")
        return redirect(url_for("book_bp.landing_page"))

    flash("Neplataný kód", "danger")
    return redirect(url_for("book_bp.landing_page"))


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@book_bp.route('/upload/source/excel/',methods=["POST"])
@check_admin
def upload_file():

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    book_type_name = request.form["book_type_name"]

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active

        book_code_column = int(request.form["book-column-code"])
        book_type = BookType.query.filter(BookType.name == book_type_name).first()

        if book_type:
            try:
                for i in range(1, sheet_obj.max_row):
                    code = sheet_obj.cell(row=i+2, column=book_code_column).value
                    if code:
                        try:
                            new_book = Book(code, book_type)
                            db_session.add(new_book)
                        except:
                            continue

                db_session.commit()
            except:
                flash("Nastala chyba pri nahrávaní", "danger")
                return redirect(url_for("book_bp.book_type_page", bt_name=book_type_name))

            flash("Učebnice z execelu boli úspešne pridané","success")
            return redirect(url_for("book_bp.book_type_page", bt_name=book_type_name))

    allow_f_string = ' / '.join(map(str, ALLOWED_EXTENSIONS))
    flash(f"Súbor nie je podporovaný. Typ súboru musí byť: {allow_f_string}","danger")
    return redirect(url_for("book_bp.book_type_page", bt_name=book_type_name))


@book_bp.route('/upload/source/excel/full/',methods=["POST"])
@check_admin
def upload_file_full():

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active

        book_code_column = int(request.form["book-column-code"])
        book_author_column = int(request.form["book-column-author"])
        book_name_column = int(request.form["book-column-name"])

        for i in range(1, sheet_obj.max_row):
            book_type_name = sheet_obj.cell(row=i+2, column=book_name_column).value
            book_type = BookType.query.filter(BookType.name == book_type_name).first()

            code = sheet_obj.cell(row=i + 2, column=book_code_column).value

            if book_type:
                if code:
                    new_book = Book(code, book_type)
                    db_session.add(new_book)
                    db_session.commit()
                    print(f"-{code}-")

            else:
                book_type_author = sheet_obj.cell(row=i+2, column=book_author_column).value
                new_book_type = BookType(book_type_name, book_type_author)

                db_session.add(new_book_type)
                db_session.commit()
                print(f"vytvorená nová kniha: Meno: {new_book_type.name} Autor: {new_book_type.author}")
                if code:
                    new_book = Book(code, new_book_type)
                    db_session.add(new_book)
                    db_session.commit()
                    print(f"-{code}-")

        db_session.commit()
        flash("Knihy boli pridané", "success")
        return redirect(url_for("book_bp.landing_page"))

    allow_f_string = ' / '.join(map(str, ALLOWED_EXTENSIONS))
    flash(f"Súbor nie je podporovaný. Typ súboru musí byť: {allow_f_string}","danger")
    return redirect(url_for("book_bp.landing_page"))

